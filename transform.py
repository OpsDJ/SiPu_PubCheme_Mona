import logging
import os
import re

import pandas as pd
import math
import os
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


def get_id_path_map(base_path):
    """递归遍历目录，建立 id -> 文件夹路径映射"""
    id_path = {}
    for root, dirs, files in os.walk(base_path):
        for d in dirs:
            if "_" in d:
                try:
                    folder_id = int(d.split("_")[0])
                    id_path[folder_id] = os.path.join(root, d)
                except ValueError:
                    continue
    return id_path


def get_sheet_info_by_id(root_dir, cat_id):
    """输入 cat_id，返回对应的 CSV 文件路径和名称（使用原 sheet_name 命名）"""
    id_path_map = get_id_path_map(root_dir)
    save_path = id_path_map.get(cat_id)

    if not save_path:
        logging.warning(f"未找到 ID={cat_id} 的分类路径")
        return None, None

    # 原 sheet_name
    current_folder = os.path.basename(save_path)
    sheet_name = "".join(re.findall(r'[\u4e00-\u9fff]+', current_folder)) or current_folder

    # CSV 文件放在上一层目录
    parent_folder = os.path.dirname(save_path)
    csv_file = os.path.join(parent_folder, f"{sheet_name}.csv")
    os.makedirs(os.path.dirname(csv_file), exist_ok=True)

    logging.info(f"生成 CSV 路径: {csv_file} | 名称: {sheet_name}")
    return csv_file


def transform_excel(csv_filepath):
    df = pd.read_csv(csv_filepath, encoding='utf-8')

    excel_file_path = csv_filepath.replace(".csv", ".xlsx")

    max_rows = 1_000_000  # 每个 sheet 放 100 万行以内
    writer = pd.ExcelWriter(excel_file_path, engine='openpyxl')

    sheet_count = math.ceil(len(df) / max_rows)

    for i in range(sheet_count):
        start = i * max_rows
        end = (i + 1) * max_rows
        df_chunk = df.iloc[start:end]
        df_chunk.to_excel(writer, sheet_name=f"Sheet{i + 1}", index=False)

    writer.close()
    print(f"转换完成：共 {sheet_count} 个sheet")





def get_img_info_by_id(root_dir, cat_id):
    """输入 cat_id，返回分类目录下的 img 文件夹路径"""
    id_path_map = get_id_path_map(root_dir)
    save_path = id_path_map.get(cat_id)

    if not save_path:
        logging.warning(f"未找到 ID={cat_id} 的分类路径")
        return None

    img_file = os.path.join(save_path, "img")
    logging.info(f"图片路径: {img_file}")
    return img_file


def insert_images_to_excel(
        excel_path,
        img_dir,
        not_found_img=r"D:\Pycharm项目\化学信息采集\not_find.png",
        image_column="C",
        header_row_index=1,
        data_row_start=2,
        adjust_row_height=True,
        img_max_width=120,
        img_max_height=120
    ):
    """
    从 Excel 中逐行读取图片文件名，如能找到图片则插入，
    若找不到则插入备用图片 not_found_img。

    参数说明见上一版本。
    """

    print("加载 Excel 文件...")
    wb = load_workbook(excel_path)
    ws = wb.active
    print(f"已加载：{excel_path}，总行数：{ws.max_row}")

    # ---- 找到“图片路径”列 ----
    img_path_col_index = None
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row_index, column=col_idx).value
        if cell_val == "图片路径":
            img_path_col_index = col_idx
            break

    if img_path_col_index is None:
        raise ValueError("未找到表头列名 '图片路径'")

    img_path_col_letter = get_column_letter(img_path_col_index)
    print(f"图片路径列定位为: {img_path_col_letter}")

    print(f"开始逐行扫描：从 {data_row_start} 到 {ws.max_row} 行")

    processed = 0

    for row in range(data_row_start, ws.max_row + 1):

        raw = ws.cell(row=row, column=img_path_col_index).value

        # 没有图片名（空行）就用 not_found 图
        if raw is None or str(raw).strip() == "" or str(raw).lower() == "nan":
            full_img_path = not_found_img
        else:
            img_name = str(raw).strip()
            full_img_path = os.path.join(img_dir, img_name)
            if not os.path.exists(full_img_path):
                full_img_path = not_found_img   # 找不到 → 使用备用图

        try:
            img = Image(full_img_path)

            # 等比缩放
            w, h = img.width, img.height
            scale = min(1.0,
                        img_max_width / w if w > 0 else 1.0,
                        img_max_height / h if h > 0 else 1.0)
            if scale < 1.0:
                img.width = int(w * scale)
                img.height = int(h * scale)

            target_cell = f"{image_column}{row}"
            ws.add_image(img, target_cell)

            # 自动调整行高
            if adjust_row_height:
                desired_height_pt = int(img.height * 0.75)
                if ws.row_dimensions[row].height is None or ws.row_dimensions[row].height < desired_height_pt:
                    ws.row_dimensions[row].height = desired_height_pt

            processed += 1

            if processed % 50 == 0:
                print(f"已插入 {processed} 张图片（到第 {row} 行）")

        except Exception as e:
            print(f"❌ 行 {row} 插图失败：{full_img_path} -> {e}")

    out_path = excel_path.replace(".xlsx", "_with_img.xlsx")
    wb.save(out_path)
    print(f"完成：共插入 {processed} 行图片。已保存：{out_path}")

if __name__ == "__main__":
    root_dir = r"./CompoundLibrary"
    cat_id = 54

    # 你已经通过 get_img_info_by_id 获取到了最终图片目录：
    img_dir = get_img_info_by_id(root_dir, cat_id)

    excel_path = r"D:\Pycharm项目\化学信息采集\CompoundLibrary\0_结构类型化合物库\2_Isoprenoids (异戊烯类化合物)\14_Terpenoids (萜类化合物)\单萜类.xlsx"
    insert_images_to_excel(excel_path, img_dir)